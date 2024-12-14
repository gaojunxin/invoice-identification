import pdfplumber
import re
import os
from pathlib import Path
import xml.etree.ElementTree as ET
# import xlwt
import openpyxl
 
def re_text(bt, text):
    m1 = re.search(bt, text)
    if m1:
        return re_block(m1.group(1))
    return ""
 
def re_block(text):
    return re.sub(r"[ ）)：]", "", text)
 

 # 读取发票
def read_invoice(workspace):
    filenames = [file.name for file in os.scandir(workspace) if file.name.endswith(".pdf")]
    row = 1
 
    result = {}
    for filename in filenames:
        
        try:
            file_path = os.path.join(workspace, filename)
            if '发票' not in filename:
                continue

            filename_stem = Path(filename).stem

            xml_file_path = file_path.replace(".pdf",".xml")
            if os.path.exists(xml_file_path):
                print("优先解析:%s.xml"%(filename_stem))
                tree = ET.parse(xml_file_path)
                root = tree.getroot()
                invoice_number = root.find('.//TaxSupervisionInfo/InvoiceNumber').text
                IssueTime = root.find('.//TaxSupervisionInfo/IssueTime').text
                taxRate = root.find('.//IssuItemInformation/TaxRate').text
                TotalTax_includedAmount = root.find('.//BasicInformation/TotalTax-includedAmount').text
                
                data = {
                        "发票代码": '',
                        "发票号码": invoice_number,
                        "开票日期": IssueTime,
                        "校验码": '',
                        "公司": '',
                        "金额": TotalTax_includedAmount,
                        "税额": '',
                        "税率": taxRate
                    }
                result[filename_stem] = (data)
                

            else:
                print("没找到【%s.xml】文件，继续解析【%s.pdf】"%(filename_stem, filename_stem))

                with pdfplumber.open(file_path) as pdf:

                    first_page = pdf.pages[0]
                    pdf_text = first_page.extract_text()
                    # print(pdf_text)
                    # 获取发票信息
                    fapiaodaima = re_text(r'发票代码.*?(\d+)', pdf_text)
                    fapiaohaoma = re_text(r'发票号码.*?(\d+)', pdf_text)
                    kaipiaoriqi = re_text(r'开票日期(.*)', pdf_text)
                    jiaoyan = re_text(r'校 验 码\s*[:：]\s*([a-zA-Z0-9 ]+)', pdf_text)
                    
    
                    heji_match = re_text(r'合.*计(.*)', pdf_text)
    #                 heji = re.split(r"￥|¥", heji_match) if heji_match else ["", ""]
                    heji = re.split(r"￥|¥", heji_match) if heji_match else ["", ""] #re.split(r" ", heji_match) #
    
                    jine = heji[1] if len(heji) > 1 else "".join(heji)
                    shuie = heji[2] if len(heji) > 2 else "".join(heji)
                    
                    # 提取公司名称
                    company_match = re.findall(r'名.*?称\s*[:：]\s*([一-龟]+)', pdf_text)
                    gongsi = re_block(company_match[-1]) if company_match else ""
    
                    data = {
                        "发票代码": fapiaodaima,
                        "发票号码": fapiaohaoma,
                        "开票日期": kaipiaoriqi,
                        "校验码": jiaoyan,
                        "公司": gongsi,
                        "金额": jine,
                        "税额": shuie,
                        "税率": '',
                    }
                    result[filename_stem] = (data)
            
 
        except Exception as exc:
            print(f"Error processing {filename}: {exc}")
    return result
 
# 读取行程单
def read_trip(workspace):
    filenames = [file.name for file in os.scandir(workspace) if file.name.endswith(".pdf")]
    row = 1

    result = []
    for filename in filenames:
        if '行程单' not in filename:
                    continue
        filename_stem = Path(filename).stem
        try:
            with pdfplumber.open(os.path.join(workspace, filename)) as pdf:
                first_page = pdf.pages[0]
 
                pdf_tables = first_page.extract_tables()
                print(filename)
                
                for table in pdf_tables:
                    for row in table[1:]:
                        row_array = []
                        if '\n' in row[0]:
                            tmp = row[0].split('\n')
                            tmp1 = tmp[1].split(' ')
                            row_array = tmp1[0:7]
                            row_array.append(tmp[0]+tmp[2])
                            row_array.append(tmp1[7])
                            # print(row_array)
                        else:
                            row_array = re.split(r'\s+', row[0])

                        if row_array:
                            result_row = {
                                "日期": row_array[3],
                                "城市": row_array[5],
                                "起点": row_array[6],
                                "终点": row_array[7],
                                "金额": row_array[8],
                                "发票文件": filename_stem.replace("电子行程单", "电子发票")
                            }
                            result.append(result_row)
 
        except Exception as exc:
            print(f"Error processing {filename}: {exc}")
     
    return result

def write(lst):
    # 创建工作簿
    wb = openpyxl.Workbook()
    # 获取活动的工作表
    ws = wb.active
    # 设置工作表的标题
    ws.title = 'Sheet 1'
    for i, name in zip(range(12), ['发票代码', '发票号码', '开票日期', "校验码", '公司', '金额', '税额', '税率', '城市', '起点', '终点', '日期','发票文件']):
        ws.cell(row=1, column=i+1, value=name)
        
    # 写入Excel
    for i, row_data in enumerate(lst, start=2):
        for j, cell_value in enumerate(row_data, start=1):
            ws.cell(row=i, column=j, value=cell_value)

    wb.save('行程信息.xlsx')


workspace = '/home/gaojunxin/图片/发票'

invoice_info = read_invoice(workspace)
print("获取发票信息如下：\n %s"%(invoice_info))

# 行程信息获取
trip_info = read_trip(workspace)
print("行程信息获取如下：\n %s"%(trip_info))


array = []
for trip_item in trip_info:
    filename_stem = trip_item['发票文件']
    invoice_item = invoice_info[filename_stem]
    
    item = [
        invoice_item['发票代码'],
        invoice_item['发票号码'],
        invoice_item['开票日期'],
        invoice_item['校验码'],
        invoice_item['公司'],
        trip_item['金额'],
        invoice_item['税额'],
        invoice_item['税率'],
        trip_item['城市'],
        trip_item['起点'],
        trip_item['终点'],
        trip_item['日期'],
        filename_stem
        ]
    array.append(item)

    write(array)